import openpyxl 
from openpyxl import Workbook, load_workbook
import re



def copy_sheet(source_wb, new_wb) -> None:
  """
  Creates the copy of source_wb into new_wb
  
  Parameters:
  source_wb (Workbook): The workbook to copy from
  new_wb (Workbook): The workbook to copy to
  """
  # Loop through all sheets in the source workbook
  for sheet_name in source_wb.sheetnames:
      # Get the source sheet
      source_sheet = source_wb[sheet_name]
      
      # Create a new sheet in the target workbook with the same name
      target_sheet = new_wb.create_sheet(title=sheet_name)
      
      # Copy the contents of the source sheet to the target sheet
      for row in source_sheet.iter_rows():
          for cell in row:
              target_sheet[cell.coordinate].value = cell.value
      
      # Copy the dimensions of the columns and rows
      for col in source_sheet.column_dimensions:
          target_sheet.column_dimensions[col] = source_sheet.column_dimensions[col]
      for row in source_sheet.row_dimensions:
          target_sheet.row_dimensions[row] = source_sheet.row_dimensions[row]

def prepare_excel_sheets() -> Workbook:
  """
  Creates a copy of the original Disney Dining Spreadsheet and returns the menu sheet. 
  This prevents modifying the original, so there is always an original copy if there were any mistakes
  
  Returns:
  menu (Workbook): The menu that will be modified
  """
  # Give the location of the file 
  path = "/Users/mokrzesik/Desktop/Michael/General/Disney/05-01-2024 Disney Dining.xlsx"

  # set the sheet names of the workbook
  sheet1 = 'Restaurant List'
  sheet2 = 'Menus'
  
  # workbook object is created 
  source_wb = load_workbook(path) 

  # Create a new workbook
  new_wb = Workbook()

  # Remove the default sheet created in the new workbook
  new_wb.remove(new_wb.active)

  # Copy the sheets into a new workbook
  copy_sheet(source_wb, new_wb)
  
  # Save the new workbook
  new_wb.save('Copy_Disney_Dining.xlsx')

  # from the active attribute 
  menu = new_wb.worksheets[1]

  # returns menu
  return menu

def load_sheets(path) -> (Workbook, Workbook):
  """
  Takes the path to the copied workbook, and returns a new workbook that will be saved somewhere else
  along with a the menu that will be modified
  
  Parameters:
  path (str): the path to the copied workbook
  
  Returns:
  menu (Worksheet): The menu that will be modified
  new_wb (Workbook): The entire workbook that is modified
  """
  # set the sheet names of the workbook
  sheet1 = 'Restaurant List'
  sheet2 = 'Menus'

  # workbook object is created
  new_wb = load_workbook(path)

  # obtains the menu sheet from the workbook
  menu = new_wb.worksheets[1]

  # returns the menu and the new workbook
  return (menu, new_wb)


# checks if the string is +- number with 2 decimal places
decimal_pattern = re.compile(r'[-+]?\d*\.\d{2}')
# checks if the string has a dollar sign
dollar_sign = re.compile(r'\$')
# checks if the string has the 'market price' phrase and ignores cases
market_price = re.compile(r'market price', re.IGNORECASE)
# checks if the strng has a number
number_pattern = re.compile(r'\d+')
# checks if the string has a date pattern
date_pattern = re.compile(r"['’‘]\d+")
# checks if the string has an age pattern
age_pattern = re.compile(r'\d+[+]')
# checks if the string has a year pattern
year_pattern = re.compile(r'\d{4}')
# checks if the string has 'oz' phrase
ounce_pattern = re.compile(r'oz')
# checks if the string has 'mL' or 'L' phrase
liquid_pattern = re.compile(r'mL|L')
# checks if the string has a double digit number
doubledigit_pattern = re.compile(r'\d{2}')
# checks if the string has multiple numbers followed by words
multiple_pattern = re.compile(r'\d+\s*\|\s*\d+\s*\|\s*\d+')
# checks if there is a single digit
singledigit_pattern = re.compile(r'\d{1}')
# checks if the string has a glass bottle pattern and ignores cases
glass_bottle_pattern = re.compile(r'glass.*bottle|bottle.*glass', re.IGNORECASE)

# Function to check if the string is a value
def value_check(string) -> bool:
  """
  Checks if a given string contains a value that meets certain criteria.

  Parameters:
  string (str): The string to be checked.

  Returns:
  bool: True if the string contains a value that meets the criteria, False otherwise.
  """  
  # checks if the string contains anything if not return false
  if string == None: return False

  # checks if there is a dollar sign, has a decimal pattern, or a multiple pattern each of with are automatically money
  if dollar_sign.search(string) or decimal_pattern.search(string) or multiple_pattern.search(string):
    return True
  
  # This is where we check for other numbers that are not money values so we can return False. 
  # We check for the following patterns:
  # - date pattern
  # - age pattern
  # - year pattern
  # - ounce pattern
  # - liquid pattern
  # - double digit pattern
  # - single digit pattern
  elif (date_pattern.search(string) or age_pattern.search(string) or year_pattern.search(string) or ounce_pattern.search(string) 
        or liquid_pattern.search(string) or doubledigit_pattern.search(string) or singledigit_pattern.search(string)):
    return False
  # after all this which should filter out every possible number that is not a money value but still a number we finally check
  # if the string has a number pattern and return True if it does
  elif number_pattern.search(string):
    return True
  # Finally return false if none of the above conditions are met which can only happen if there is no number in the string which
  # should not happen
  else:
    return False

# Function to process the string down to just the numbers
def process_string(input_string):
  """
  Process a string by removing certain patterns and extracting numbers.

  Parameters:
  input_string (str): The input string to be processed.

  Returns:
  str: The processed string containing extracted numbers separated by hyphens if necessary.
  """
  # Remove dollar signs
  string = re.sub(r'\$', '', input_string)

  # Search for the pattern 'Bottle' followed by a number with optional decimal places
  match = re.search(r'Glass\s*(\d+(?:\.\d+)?)\s*\|\s*Bottle\s*(\d+(?:\.\d+)?)|Bottle\s*(\d+(?:\.\d+)?)\s*\|\s*Glass\s*(\d+(?:\.\d+)?)',
                       input_string,
                       re.IGNORECASE)
  if match:
    glass_price = match.group(1)
    bottle_price = match.group(2)
    return [glass_price, bottle_price]
    
  # Remove the word 'beverage' or 'beverages' (case-insensitive)
  string = re.sub(r'beverages?', '', string, flags=re.IGNORECASE)
  
  # Remove the word 'child' followed by an optional hyphen and number
  string = re.sub(r'child\s*\(?(\d+-)?\d\)?', '', string, flags=re.IGNORECASE)
  
  # Remove the word 'ages' followed by an optional hyphen and number
  string = re.sub(r'ages\s*\(?(\d+-)?\d\)?', '', string, flags=re.IGNORECASE)
  
  # Remove any pattern of a word followed by a 4-digit number followed by a word
  string = re.sub(r'[a-zA-Z]+\s*\d{4}\s*[a-zA-Z]+', '', string)

  # Extract numbers with optional decimal places and hyphen if necessary
  numbers = re.findall(r'\d+(?:\.\d+)?-?\d+(?:\.\d+)?', string)

  # Join the numbers into a single string separated by hyphens
  result = '-'.join(numbers)
  
  # Return the processed string
  return result

def user_Input(menu, i) -> str:
  """
  Ask the user what result they would like to be based on all the information given.
  
  Parameters:
  menu (Worksheet): The worksheet containing the menu data.
  i (int): The row currently in the sheet.
  
  Returns:
  result (str): The result to output
  """
  # Asks all the questions
  result = input(f"Here is the inputs;\n"
                  f"Title: {menu[f'A{i+1}'].value}\n"
                  f"Name: {menu[f'C{i+1}'].value}\n"
                  f"mealPeriods.groups.name: {menu[f'D{i+1}'].value}\n"
                  f"mealPeriods.groups.type: {menu[f'E{i+1}'].value}\n"
                  f"mealPeriods.name: {menu[f'F{i+1}'].value}\n"
                  f"mealPeriods.label: {menu[f'G{i+1}'].value}\n"
                  f"mealPeriods.experience: {menu[f'H{i+1}'].value}\n"
                  f"mealPeriods.serviceStyle: {menu[f'I{i+1}'].value}\n"
                  f"description: {menu[f'J{i+1}'].value}\n")
  # retuns what the user wants
  return result
                
def glass_bottle_row_duplicate(menu, i, cellA_value, result) -> None:
  """
  Creates two lines for the glass and bottle prices.
  
  Parameters:
  menu (Worksheet): The worksheet containing the menu data.
  i (int): The row currently in the sheet.
  cellA_value (str): Contents of the original value in cellA
  result (list): The list of glass and bottle prices
  """
  menu[f'A{i+1}'] = cellA_value + f' - Glass'
  menu[f'B{i+1}'] = result[0]
  menu[f'L{i+1}'] = f'Mod: J'
  menu.insert_rows(i+2)
  menu[f'A{i+2}'] = cellA_value + f' - Bottle'
  menu[f'B{i+2}'] = result[1]
  menu[f'C{i+2}'] = menu[f'C{i+1}'].value
  menu[f'D{i+2}'] = menu[f'D{i+1}'].value
  menu[f'E{i+2}'] = menu[f'E{i+1}'].value
  menu[f'F{i+2}'] = menu[f'F{i+1}'].value
  menu[f'G{i+2}'] = menu[f'G{i+1}'].value
  menu[f'H{i+2}'] = menu[f'H{i+1}'].value
  menu[f'I{i+2}'] = menu[f'I{i+1}'].value
  menu[f'J{i+2}'] = menu[f'J{i+1}'].value
  menu[f'K{i+2}'] = menu[f'K{i+1}'].value
  menu[f'L{i+2}'] = menu[f'J{i+1}'].value